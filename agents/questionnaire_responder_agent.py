from __future__ import annotations

import json
from pathlib import Path

from agents.base_agent import BaseAgent, _strip_fences
from core.models import QuestionnaireAnswer
from core.vector_store import VectorStore

SYSTEM_PROMPT = """You are a Head of GRC at a SaaS company answering a security questionnaire on behalf of your organization.

Rules:
1. Use ONLY the provided CCF/policy context to answer. Do not hallucinate.
2. If a question has no grounding in the context, set confidence to "NEEDS_HUMAN_REVIEW" and explain why.
3. Maintain a professional, confident tone appropriate for enterprise security reviews.
4. Reference the exact control IDs or policy sections you are drawing from.
5. Be precise — evaluators are security professionals.

Respond with valid JSON only (no markdown fences):
{
  "question": "...",
  "answer": "...",
  "confidence": "HIGH" | "MEDIUM" | "NEEDS_HUMAN_REVIEW",
  "source_references": ["CCF control ID or policy section name"]
}"""


class QuestionnaireResponderAgent(BaseAgent):
    def __init__(self, vector_store: VectorStore | None = None) -> None:
        super().__init__(system_prompt=SYSTEM_PROMPT)
        if vector_store is not None:
            self.vector_store: VectorStore | None = vector_store
        else:
            try:
                self.vector_store = VectorStore()
            except RuntimeError:
                self.vector_store = None  # chromadb not installed

    def answer_question(self, question: str) -> QuestionnaireAnswer:
        if self.vector_store is None:
            context = "No knowledge base available — chromadb is not installed."
        else:
            chunks = self.vector_store.query(question, n_results=3)
            context = "\n\n---\n\n".join(chunks) if chunks else "No relevant context found."
        prompt = (
            f"Question: {question}\n\n"
            f"Relevant CCF/Policy context:\n{context}\n\n"
            "Return valid JSON matching the QuestionnaireAnswer schema."
        )
        raw = self.run(prompt)
        try:
            data = json.loads(_strip_fences(raw))
            return QuestionnaireAnswer(**data)
        except Exception:
            return QuestionnaireAnswer(
                question=question,
                answer="Unable to parse agent response.",
                confidence="NEEDS_HUMAN_REVIEW",
                source_references=[],
            )

    def answer_questionnaire(
        self, questions: list[str]
    ) -> list[QuestionnaireAnswer]:
        return [self.answer_question(q) for q in questions]

    def export_to_excel(self, answers: list[QuestionnaireAnswer], output_path: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Security Questionnaire"

            headers = ["Question", "Answer", "Confidence", "Source References"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)

            CONFIDENCE_COLORS = {
                "HIGH": "C6EFCE",
                "MEDIUM": "FFEB9C",
                "NEEDS_HUMAN_REVIEW": "FFC7CE",
            }

            for row_idx, ans in enumerate(answers, 2):
                ws.cell(row=row_idx, column=1, value=ans.question)
                ws.cell(row=row_idx, column=2, value=ans.answer)
                conf_cell = ws.cell(row=row_idx, column=3, value=ans.confidence)
                fill_color = CONFIDENCE_COLORS.get(ans.confidence, "FFFFFF")
                conf_cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                ws.cell(row=row_idx, column=4, value=", ".join(ans.source_references))

            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 80
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 40

            path = Path(output_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(path))
            return str(path)
        except ImportError:
            return "openpyxl not installed — cannot export to Excel."
